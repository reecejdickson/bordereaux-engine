"""
Full bordereaux validator with Excel support and 10 validation rules.
"""
from typing import List, Dict, Any, Optional, Tuple
import csv
import io
import re
from datetime import datetime
from collections import defaultdict

# Try to import openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


def get_column_letter(col_idx: int) -> str:
    """Convert column index to Excel letter (0=A, 1=B, etc.)"""
    result = ""
    while col_idx >= 0:
        result = chr(col_idx % 26 + 65) + result
        col_idx = col_idx // 26 - 1
    return result


def parse_csv(content: bytes) -> Tuple[List[str], List[List[Any]], str]:
    """Parse CSV file, return headers, rows, sheet_name"""
    text = content.decode('utf-8', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [], "Sheet1"
    return rows[0], rows[1:], "Sheet1"


def parse_excel(content: bytes) -> Tuple[List[str], List[List[Any]], str]:
    """Parse Excel file, return headers, rows, sheet_name"""
    if not EXCEL_SUPPORT:
        raise ValueError("Excel support not available")
    
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active
    sheet_name = sheet.title
    
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    
    if not rows:
        return [], [], sheet_name
    
    # Filter out completely empty rows
    rows = [r for r in rows if any(cell is not None for cell in r)]
    
    if not rows:
        return [], [], sheet_name
    
    return [str(h) if h else "" for h in rows[0]], rows[1:], sheet_name


def validate_file(content: bytes, filename: str, record_type: str) -> List[Dict]:
    """
    Validate a bordereaux file and return list of issues.
    Supports both CSV and Excel files.
    Implements 10 validation rules.
    """
    issues = []
    
    # Determine file type and parse
    try:
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        
        if ext in ['xlsx', 'xls', 'xlsm']:
            if not EXCEL_SUPPORT:
                issues.append(create_issue("PARSE001", "P0", "Excel not supported",
                    "Install openpyxl for Excel support", "N/A", filename, "CSV file"))
                return issues
            headers, data_rows, sheet_name = parse_excel(content)
        else:
            headers, data_rows, sheet_name = parse_csv(content)
        
        if not headers:
            issues.append(create_issue("FILE001", "P0", "Empty file",
                "File contains no headers", "A1", "Empty", "Header row"))
            return issues
        
        if not data_rows:
            issues.append(create_issue("FILE002", "P0", "No data rows",
                "File contains headers but no data", "A2", "Empty", "Data rows"))
            return issues
        
    except Exception as e:
        issues.append(create_issue("PARSE001", "P0", "Could not parse file",
            f"Error: {str(e)[:200]}", "N/A", filename, "Valid file"))
        return issues
    
    # Normalize headers for matching
    headers_lower = [str(h).lower().strip() for h in headers]
    
    # Find column indices for key fields
    col_map = find_columns(headers_lower, record_type)
    
    # Run all validation rules
    issues.extend(rule_req001_required_fields(headers_lower, record_type, col_map))
    issues.extend(rule_date001_date_logic(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_ccy001_currency(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_num001_negative_premium(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_num002_share_percentage(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_dup001_duplicates(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_calc001_premium_calculation(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_type001_data_types(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_rec001_premium_reconciliation(data_rows, headers, col_map, sheet_name))
    issues.extend(rule_req002_empty_cells(data_rows, headers, col_map, sheet_name))
    
    return issues


def find_columns(headers: List[str], record_type: str) -> Dict[str, int]:
    """Find column indices for known field patterns"""
    patterns = {
        'policy': ['policy', 'policy_ref', 'policy_number', 'pol_no', 'reference'],
        'claim': ['claim', 'claim_ref', 'claim_number', 'clm_no'],
        'insured': ['insured', 'insured_name', 'client', 'policyholder', 'assured'],
        'inception': ['inception', 'inception_date', 'incept', 'start_date', 'effective'],
        'expiry': ['expiry', 'expiry_date', 'expire', 'end_date', 'termination'],
        'gross_premium': ['gross_premium', 'gross', 'premium', 'gwp', 'gross_written'],
        'net_premium': ['net_premium', 'net', 'nwp', 'net_written'],
        'commission': ['commission', 'broker_commission', 'brokerage', 'comm'],
        'currency': ['currency', 'ccy', 'curr'],
        'share': ['share', 'our_share', 'participation', 'line', 'signing'],
        'sum_insured': ['sum_insured', 'tsi', 'total_sum', 'limit', 'tiv'],
        'transaction_type': ['transaction', 'trans_type', 'type', 'movement'],
        'loss_date': ['loss_date', 'date_of_loss', 'dol', 'occurrence_date'],
        'paid': ['paid', 'paid_amount', 'payment', 'paid_this_period'],
        'reserve': ['reserve', 'outstanding', 'os_reserve', 'incurred'],
    }
    
    col_map = {}
    for field, field_patterns in patterns.items():
        for idx, header in enumerate(headers):
            for pattern in field_patterns:
                if pattern in header:
                    col_map[field] = idx
                    break
            if field in col_map:
                break
    
    return col_map


def create_issue(rule_id: str, severity: str, title: str, description: str,
                 cell_ref: str, actual: str, expected: str, row_idx: int = None,
                 calculation: Dict = None) -> Dict:
    """Create a standardized issue dict"""
    issue = {
        "rule_id": rule_id,
        "severity": severity,
        "title": title,
        "description": description,
        "cell_reference": cell_ref,
        "actual_value": str(actual)[:500] if actual else None,
        "expected_value": str(expected)[:500] if expected else None,
        "row_index": row_idx,
    }
    if calculation:
        issue["calculation"] = calculation
    return issue


def get_cell_value(row: List, col_idx: int) -> Any:
    """Safely get cell value"""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


# =============================================================================
# VALIDATION RULES
# =============================================================================

def rule_req001_required_fields(headers: List[str], record_type: str, col_map: Dict) -> List[Dict]:
    """REQ001: Check for required fields based on record type"""
    issues = []
    
    if record_type == "premium":
        required = ['policy', 'gross_premium', 'insured', 'inception']
    else:  # claims
        required = ['claim', 'policy', 'loss_date']
    
    for field in required:
        if field not in col_map:
            issues.append(create_issue(
                "REQ001", "P0", f"Missing required field: {field}",
                f"Could not find a column for '{field}'",
                "Row 1", ", ".join(headers[:5]), f"Column containing '{field}'"
            ))
    
    return issues


def rule_date001_date_logic(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """DATE001: Inception date must be before or equal to expiry date"""
    issues = []
    
    inception_idx = col_map.get('inception')
    expiry_idx = col_map.get('expiry')
    
    if inception_idx is None or expiry_idx is None:
        return issues
    
    for row_idx, row in enumerate(rows, start=2):
        inception = get_cell_value(row, inception_idx)
        expiry = get_cell_value(row, expiry_idx)
        
        if not inception or not expiry:
            continue
        
        try:
            # Handle datetime objects from Excel
            if isinstance(inception, datetime):
                inc_date = inception
            else:
                inc_date = parse_date(str(inception))
            
            if isinstance(expiry, datetime):
                exp_date = expiry
            else:
                exp_date = parse_date(str(expiry))
            
            if inc_date and exp_date and inc_date > exp_date:
                col_letter = get_column_letter(inception_idx)
                issues.append(create_issue(
                    "DATE001", "P0", f"Inception after expiry (Row {row_idx})",
                    "Inception date is after expiry date",
                    f"{col_letter}{row_idx}",
                    f"Inception: {inception}, Expiry: {expiry}",
                    "Inception ≤ Expiry",
                    row_idx
                ))
        except:
            pass
    
    return issues


def rule_ccy001_currency(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """CCY001: Currency codes should be valid ISO 4217"""
    issues = []
    
    ccy_idx = col_map.get('currency')
    if ccy_idx is None:
        return issues
    
    valid_currencies = {'USD', 'GBP', 'EUR', 'CAD', 'AUD', 'JPY', 'CHF', 'NZD', 'ZAR', 
                       'SGD', 'HKD', 'NOK', 'SEK', 'DKK', 'MXN', 'BRL', 'INR', 'CNY'}
    
    for row_idx, row in enumerate(rows, start=2):
        ccy = get_cell_value(row, ccy_idx)
        if not ccy:
            continue
        
        ccy_str = str(ccy).strip().upper()
        if ccy_str and ccy_str not in valid_currencies:
            col_letter = get_column_letter(ccy_idx)
            issues.append(create_issue(
                "CCY001", "P1", f"Invalid currency code (Row {row_idx})",
                f"'{ccy_str}' is not a recognized ISO currency code",
                f"{col_letter}{row_idx}",
                ccy_str, "Valid ISO 4217 code (USD, GBP, EUR, etc.)",
                row_idx
            ))
    
    return issues


def rule_num001_negative_premium(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """NUM001: Negative premiums should only appear with cancellation transactions"""
    issues = []
    
    premium_idx = col_map.get('gross_premium')
    trans_idx = col_map.get('transaction_type')
    
    if premium_idx is None:
        return issues
    
    cancel_keywords = ['cancel', 'can', 'return', 'refund', 'reversal', 'void']
    
    for row_idx, row in enumerate(rows, start=2):
        premium = get_cell_value(row, premium_idx)
        trans_type = get_cell_value(row, trans_idx) if trans_idx else None
        
        if not premium:
            continue
        
        try:
            premium_val = parse_number(premium)
            if premium_val is not None and premium_val < 0:
                # Check if it's a cancellation
                is_cancel = False
                if trans_type:
                    trans_str = str(trans_type).lower()
                    is_cancel = any(kw in trans_str for kw in cancel_keywords)
                
                if not is_cancel:
                    col_letter = get_column_letter(premium_idx)
                    issues.append(create_issue(
                        "NUM001", "P1", f"Negative premium without cancellation (Row {row_idx})",
                        "Premium is negative but transaction type doesn't indicate cancellation",
                        f"{col_letter}{row_idx}",
                        str(premium_val), "Positive value or cancellation transaction",
                        row_idx
                    ))
        except:
            pass
    
    return issues


def rule_num002_share_percentage(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """NUM002: Share percentage must be between 0 and 100"""
    issues = []
    
    share_idx = col_map.get('share')
    if share_idx is None:
        return issues
    
    for row_idx, row in enumerate(rows, start=2):
        share = get_cell_value(row, share_idx)
        if not share:
            continue
        
        try:
            share_val = parse_number(share)
            if share_val is not None:
                # Normalize to percentage
                if share_val > 1 and share_val <= 100:
                    share_pct = share_val
                elif share_val >= 0 and share_val <= 1:
                    share_pct = share_val * 100
                else:
                    share_pct = share_val
                
                if share_pct < 0 or share_pct > 100:
                    col_letter = get_column_letter(share_idx)
                    issues.append(create_issue(
                        "NUM002", "P1", f"Invalid share percentage (Row {row_idx})",
                        f"Share {share_val} is outside valid range",
                        f"{col_letter}{row_idx}",
                        str(share_val), "0-100%",
                        row_idx
                    ))
        except:
            pass
    
    return issues


def rule_dup001_duplicates(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """DUP001: Check for duplicate policy/claim references"""
    issues = []
    
    ref_idx = col_map.get('policy') or col_map.get('claim')
    if ref_idx is None:
        return issues
    
    seen = {}
    for row_idx, row in enumerate(rows, start=2):
        ref = get_cell_value(row, ref_idx)
        if not ref:
            continue
        
        ref_str = str(ref).strip()
        if ref_str in seen:
            col_letter = get_column_letter(ref_idx)
            issues.append(create_issue(
                "DUP001", "P1", f"Duplicate reference: {ref_str}",
                f"Reference '{ref_str}' appears in rows {seen[ref_str]} and {row_idx}",
                f"{col_letter}{row_idx}",
                ref_str, "Unique reference",
                row_idx
            ))
        else:
            seen[ref_str] = row_idx
    
    return issues


def rule_calc001_premium_calculation(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """CALC001: Net premium should equal gross premium minus commission"""
    issues = []
    
    gross_idx = col_map.get('gross_premium')
    net_idx = col_map.get('net_premium')
    comm_idx = col_map.get('commission')
    
    if not all([gross_idx is not None, net_idx is not None, comm_idx is not None]):
        return issues
    
    for row_idx, row in enumerate(rows, start=2):
        gross = parse_number(get_cell_value(row, gross_idx))
        net = parse_number(get_cell_value(row, net_idx))
        comm = parse_number(get_cell_value(row, comm_idx))
        
        if gross is None or net is None or comm is None:
            continue
        
        expected_net = gross - comm
        diff = abs(expected_net - net)
        
        if diff > 0.01:  # Allow small rounding
            col_letter = get_column_letter(net_idx)
            issues.append(create_issue(
                "CALC001", "P1", f"Premium calculation mismatch (Row {row_idx})",
                f"Gross ({gross:.2f}) - Commission ({comm:.2f}) = {expected_net:.2f}, but Net is {net:.2f}",
                f"{col_letter}{row_idx}",
                f"{net:.2f}", f"{expected_net:.2f}",
                row_idx,
                {"gross": gross, "commission": comm, "expected_net": expected_net, "actual_net": net}
            ))
    
    return issues


def rule_type001_data_types(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """TYPE001: Validate data types (dates should be dates, numbers should be numbers)"""
    issues = []
    
    date_fields = ['inception', 'expiry', 'loss_date']
    number_fields = ['gross_premium', 'net_premium', 'commission', 'share', 'sum_insured', 'paid', 'reserve']
    
    for row_idx, row in enumerate(rows, start=2):
        # Check date fields
        for field in date_fields:
            idx = col_map.get(field)
            if idx is None:
                continue
            val = get_cell_value(row, idx)
            if val and not isinstance(val, datetime):
                if not is_valid_date_string(str(val)):
                    col_letter = get_column_letter(idx)
                    issues.append(create_issue(
                        "TYPE001", "P2", f"Invalid date format (Row {row_idx})",
                        f"'{field}' value '{val}' is not a valid date",
                        f"{col_letter}{row_idx}",
                        str(val), "Date format (YYYY-MM-DD, DD/MM/YYYY, etc.)",
                        row_idx
                    ))
        
        # Check number fields
        for field in number_fields:
            idx = col_map.get(field)
            if idx is None:
                continue
            val = get_cell_value(row, idx)
            if val is not None and val != '':
                if not isinstance(val, (int, float)) and parse_number(val) is None:
                    col_letter = get_column_letter(idx)
                    issues.append(create_issue(
                        "TYPE001", "P2", f"Invalid number format (Row {row_idx})",
                        f"'{field}' value '{val}' is not a valid number",
                        f"{col_letter}{row_idx}",
                        str(val), "Numeric value",
                        row_idx
                    ))
    
    return issues


def rule_rec001_premium_reconciliation(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """REC001: Sum of premiums should match any stated total"""
    issues = []
    
    premium_idx = col_map.get('gross_premium')
    if premium_idx is None:
        return issues
    
    # Calculate sum
    total = 0
    count = 0
    for row in rows:
        val = parse_number(get_cell_value(row, premium_idx))
        if val is not None:
            total += val
            count += 1
    
    # Look for a "Total" row (last rows often have totals)
    for row_idx, row in enumerate(rows[-5:], start=len(rows)-4):
        first_cell = str(get_cell_value(row, 0) or '').lower()
        if 'total' in first_cell or 'sum' in first_cell:
            stated = parse_number(get_cell_value(row, premium_idx))
            if stated is not None and abs(stated - total) > 0.01:
                issues.append(create_issue(
                    "REC001", "P0", "Premium reconciliation failure",
                    f"Sum of premiums ({total:.2f}) doesn't match stated total ({stated:.2f})",
                    f"{get_column_letter(premium_idx)}{row_idx+2}",
                    f"{stated:.2f}", f"{total:.2f}",
                    row_idx + 2,
                    {"calculated_total": total, "stated_total": stated, "difference": total - stated, "row_count": count}
                ))
    
    return issues


def rule_req002_empty_cells(rows: List[List], headers: List[str], col_map: Dict, sheet: str) -> List[Dict]:
    """REQ002: Check for empty cells in key columns"""
    issues = []
    
    # Only check first 50 rows and first 10 columns to avoid noise
    key_cols = list(col_map.values())[:10]
    
    for row_idx, row in enumerate(rows[:50], start=2):
        for col_idx in key_cols:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or (isinstance(val, str) and not val.strip()):
                col_letter = get_column_letter(col_idx)
                col_name = headers[col_idx] if col_idx < len(headers) else f"Column {col_idx+1}"
                issues.append(create_issue(
                    "REQ002", "P2", f"Empty cell at {col_letter}{row_idx}",
                    f"Cell in '{col_name}' column is empty",
                    f"{col_letter}{row_idx}",
                    "(empty)", "Value required",
                    row_idx
                ))
    
    return issues


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_number(val: Any) -> Optional[float]:
    """Parse a value as a number, handling currency symbols and commas"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip()
        s = re.sub(r'[£$€,\s]', '', s)
        return float(s)
    except:
        return None


def parse_date(val: str) -> Optional[datetime]:
    """Try to parse a string as a date"""
    formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']
    for fmt in formats:
        try:
            return datetime.strptime(val.strip(), fmt)
        except:
            continue
    return None


def is_valid_date_string(val: str) -> bool:
    """Check if a string looks like a date"""
    return parse_date(val) is not None
